"""Export skills一覧.md to formatted Excel workbook."""
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MD_PATH = ROOT / "skills一覧.md"
OUT_PATH = ROOT / "skills一覧.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
META_FONT = Font(name="Arial", size=10, bold=True)
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def parse_markdown(text: str):
    sections = []
    current_section = None
    current_subsection = None
    in_table = False
    table_headers = None

    for line in text.splitlines():
        if m := re.match(r"^## (\d+)\.\s+(.+)$", line):
            current_section = {"no": int(m.group(1)), "title": m.group(2), "subsections": [], "rows": []}
            current_subsection = None
            sections.append(current_section)
            in_table = False
            table_headers = None
            continue

        if m := re.match(r"^### (.+)$", line):
            if current_section is not None:
                current_subsection = m.group(1)
                current_section["subsections"].append({"title": current_subsection, "rows": []})
            in_table = False
            table_headers = None
            continue

        if line.strip().startswith("|") and "---" not in line:
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            if not in_table:
                table_headers = cells
                in_table = True
                continue
            if len(cells) < 3:
                continue
            cmd, desc, usage = cells[0], cells[1], cells[2]
            manual = "手動のみ" in cmd
            cmd_clean = cmd.replace("（手動のみ）", "").strip().strip("`")
            row = {
                "command": cmd_clean,
                "description": desc,
                "usage": usage,
                "manual_only": manual,
                "subsection": current_subsection or "",
            }
            if current_section is not None:
                if current_subsection and current_section["subsections"]:
                    current_section["subsections"][-1]["rows"].append(row)
                else:
                    current_section["rows"].append(row)
            continue

        if not line.strip().startswith("|"):
            in_table = False
            table_headers = None

    return sections


def style_header_row(ws, row: int, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_data_rows(ws, start_row: int, rows: list[dict], category: str, subcategory: str = ""):
    headers = ["No.", "カテゴリ", "サブカテゴリ", "コマンド", "説明", "使いどころ", "手動のみ"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    style_header_row(ws, start_row, len(headers))

    r = start_row + 1
    for i, row in enumerate(rows, 1):
        values = [
            i,
            category,
            subcategory,
            row["command"],
            row["description"],
            row["usage"],
            "○" if row["manual_only"] else "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            if (r - start_row) % 2 == 0:
                cell.fill = ALT_FILL
        r += 1
    return r


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def build_overview_sheet(wb: Workbook, md_text: str):
    ws = wb.active
    ws.title = "概要"
    meta_lines = []
    for line in md_text.splitlines():
        if line.startswith("## "):
            break
        if line.strip() and not line.startswith("# ") and not line.startswith("---"):
            meta_lines.append(line.strip())

    ws["A1"] = "Cursor グローバルスキル一覧"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:F1")

    r = 3
    for line in meta_lines:
        if line.startswith("**") and line.endswith("**"):
            ws.cell(row=r, column=1, value=line.strip("*"))
            ws.cell(row=r, column=1).font = META_FONT
            r += 1
        elif line.startswith("- **"):
            m = re.match(r"- \*\*(.+?)\*\*:\s*(.+)", line)
            if m:
                ws.cell(row=r, column=1, value=m.group(1))
                ws.cell(row=r, column=2, value=m.group(2))
                ws.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10)
                ws.cell(row=r, column=2).font = BODY_FONT
                ws.cell(row=r, column=2).alignment = WRAP
                r += 1
        elif re.match(r"^\d+\.", line):
            ws.cell(row=r, column=1, value=line)
            ws.cell(row=r, column=1).font = BODY_FONT
            r += 1

    r += 1
    ws.cell(row=r, column=1, value="カテゴリ一覧")
    ws.cell(row=r, column=1).font = META_FONT
    r += 1
    for line in md_text.splitlines():
        if m := re.match(r"^\d+\.\s+\[(.+?)\]", line):
            ws.cell(row=r, column=1, value=m.group(1))
            ws.cell(row=r, column=1).font = BODY_FONT
            r += 1

    set_col_widths(ws, {"A": 28, "B": 72, "C": 20, "D": 20, "E": 20, "F": 20})
    ws.freeze_panes = "A3"


def build_master_sheet(wb: Workbook, sections: list):
    ws = wb.create_sheet("スキル一覧")
    all_rows = []
    for sec in sections:
        cat = f"{sec['no']}. {sec['title']}"
        for row in sec["rows"]:
            all_rows.append({**row, "category": cat})
        for sub in sec.get("subsections", []):
            for row in sub["rows"]:
                all_rows.append({**row, "category": cat, "subsection": sub["title"]})

    headers = ["No.", "カテゴリ", "サブカテゴリ", "コマンド", "説明", "使いどころ", "手動のみ"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    for i, row in enumerate(all_rows, 1):
        r = i + 1
        values = [
            i,
            row["category"],
            row.get("subsection", ""),
            row["command"],
            row["description"],
            row["usage"],
            "○" if row["manual_only"] else "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_rows) + 1}"
    ws.freeze_panes = "A2"
    set_col_widths(ws, {"A": 6, "B": 36, "C": 18, "D": 28, "E": 52, "F": 44, "G": 10})
    ws.row_dimensions[1].height = 22


def sanitize_sheet_name(name: str) -> str:
    for ch in r'\/?*[]':
        name = name.replace(ch, "・")
    return name[:31]


def build_category_sheets(wb: Workbook, sections: list):
    for sec in sections:
        short = re.sub(r"[（(].+[）)]", "", sec["title"]).strip()[:20]
        name = sanitize_sheet_name(f"{sec['no']}_{short}")
        ws = wb.create_sheet(name)

        ws["A1"] = f"{sec['no']}. {sec['title']}"
        ws["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E79")
        ws.merge_cells("A1:G1")

        next_row = 3
        if sec["rows"]:
            next_row = write_data_rows(ws, next_row, sec["rows"], sec["title"])
            next_row += 2

        for sub in sec.get("subsections", []):
            if not sub["rows"]:
                continue
            ws.cell(row=next_row, column=1, value=sub["title"])
            ws.cell(row=next_row, column=1).font = META_FONT
            ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=7)
            next_row += 1
            next_row = write_data_rows(ws, next_row, sub["rows"], sec["title"], sub["title"])
            next_row += 2

        set_col_widths(ws, {"A": 6, "B": 28, "C": 18, "D": 28, "E": 52, "F": 44, "G": 10})
        ws.freeze_panes = "A3"


def main():
    md_text = MD_PATH.read_text(encoding="utf-8")
    sections = parse_markdown(md_text)

    wb = Workbook()
    build_overview_sheet(wb, md_text)
    build_master_sheet(wb, sections)
    build_category_sheets(wb, sections)

    wb.save(OUT_PATH)
    total = sum(
        len(s["rows"]) + sum(len(sub["rows"]) for sub in s.get("subsections", []))
        for s in sections
    )
    print(f"Saved: {OUT_PATH}")
    print(f"Sections: {len(sections)}, Rows: {total}")


if __name__ == "__main__":
    main()
